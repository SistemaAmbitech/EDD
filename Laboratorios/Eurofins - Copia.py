import pandas as pd
from openpyxl import Workbook

def main():
    # Caminho do arquivo Excel
    caminho = r"C:\Users\henrique.canhadas\OneDrive - Servmar Ambientais\Documentos\Codigos\GitHub Tecnologia\ProjetoCanhadas\0-Excel de Exemplos-0\Daodos EUROFINS.xlsx"
    tabela = pd.read_excel(caminho)

    # URL do arquivo Excel no GitHub (caminho local)
    url = r"C:\Users\henrique.canhadas\OneDrive - Servmar Ambientais\Documentos\Codigos\GitHub Tecnologia\ProjetoCanhadas\Tabelas Consulta\Banco de Dados\banco de dados - EUROFINS.xlsx"
    tabela_bd = pd.read_excel(url)
    colunas_merge = ["Análise"]

    tabela_merge = pd.merge(tabela, tabela_bd, how='left', on=colunas_merge)
    data_frame = tabela_merge

    resultado_final = pd.DataFrame()
    workbook = Workbook()

    for descricao_metodo in data_frame['Tipo de Método'].unique():
        filtro = (data_frame['Tipo de Método'] == descricao_metodo)
        tabela_filtrada = data_frame[filtro]

        if not tabela_filtrada.empty:
            aba_titulo = descricao_metodo[:31] if len(descricao_metodo) > 31 else descricao_metodo
            aba = workbook.create_sheet(title=aba_titulo)
            colunas_desejadas = ["Identificação da Amostra", "Data da Situação", "CASNUMBER", "Análise", "Resultado", "Unidade", "Tipo de Método"]
            aba.append(colunas_desejadas)

            for linha in tabela_filtrada[colunas_desejadas].itertuples(index=False):
                aba.append(linha)

    # Remover a primeira aba vazia criada automaticamente
    workbook.remove(workbook.worksheets[0])

    # Alterar cabeçalhos de todas as abas
    new_headers = ["SAMPLENAME", "SAMPDATE", "CASNUMBER_x", "ANALYTE", "Result", "UNITS", "Description"]
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(new_headers)):
            for cell, new_header in zip(row, new_headers):
                cell.value = new_header

    # Salva a tabela merge em "Resultado_Final_Com_Abas.xlsx"
    workbook.save(r"C:\Users\henrique.canhadas\OneDrive - Servmar Ambientais\Documentos\Codigos\GitHub Tecnologia\ProjetoCanhadas\0-Excel de Exemplos-0\EUROFINSTESTEADO.xlsx")

if __name__ == "__main__":
    main()
